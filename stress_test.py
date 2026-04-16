import pandas as pd
import numpy as np
import pickle

# ── Chargement des données de DS1 ──────────────────────────
df = pd.read_csv("clients.csv")

with open("model.pkl", "rb") as f:
    model, scaler = pickle.load(f)

print(f"✅ {len(df)} clients chargés")
print(df.head(3))
# ── Calcul des probabilités de défaut via DS1 ──────────────
features = df.drop("defaut", axis=1)
features_sc = scaler.transform(features)

df["proba_defaut"] = model.predict_proba(features_sc)[:, 1]

print("\n✅ Probabilités calculées")
print(df[["montant_demande", "incidents_passes", "proba_defaut"]].head(10))

# ── Application des 3 scénarios de stress ──────────────────
LGD = 0.45  # Loss Given Default — standard Bâle II

df["pd_normal"]  = df["proba_defaut"]
df["pd_modere"]  = (df["proba_defaut"] * 1.5).clip(0, 1)
df["pd_severe"]  = (df["proba_defaut"] * 2.0).clip(0, 1)

df["el_normal"]  = df["pd_normal"] * df["montant_demande"] * LGD
df["el_modere"]  = df["pd_modere"] * df["montant_demande"] * LGD
df["el_severe"]  = df["pd_severe"] * df["montant_demande"] * LGD

print("\n✅ Scénarios calculés")
print(df[["montant_demande", "pd_normal", "pd_severe", "el_normal", "el_severe"]].head(5))

# ── Agrégation par secteur ──────────────────────────────────
secteurs = ["Immobilier", "Commerce", "Industrie", "Services", "Agriculture"]
np.random.seed(42)
df["secteur"] = np.random.choice(secteurs, len(df))

resume = df.groupby("secteur").agg(
    nb_clients    = ("montant_demande", "count"),
    el_normal     = ("el_normal", "sum"),
    el_modere     = ("el_modere", "sum"),
    el_severe     = ("el_severe", "sum")
).round(0)

print("\n✅ Pertes par secteur (en €)")
print(resume)

# ── Rapport Excel automatique ───────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws.title = "Stress Test"

# Styles
NAVY  = "1B3A6B"
BLUE  = "D6E4F0"
WHITE = "FFFFFF"

def style(cell, bg=WHITE, bold=False, color="000000"):
    cell.font      = Font(bold=bold, color=color, name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    s = Side(style="thin", color="CCCCCC")
    cell.border    = Border(left=s, right=s, top=s, bottom=s)

# Titre
ws.merge_cells("A1:E1")
ws["A1"] = "🏦 CreditSys — Rapport Stress Test Portefeuille"
style(ws["A1"], bg=NAVY, bold=True, color=WHITE)
ws.row_dimensions[1].height = 30

# En-têtes
headers = ["Secteur", "Nb Clients", "Perte Normale (€)", "Perte Modérée (€)", "Perte Sévère (€)"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    style(c, bg=NAVY, bold=True, color=WHITE)
    ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 20

# Données
for row_idx, (secteur, row) in enumerate(resume.iterrows(), 3):
    bg = "F2F3F4" if row_idx % 2 == 0 else WHITE
    ws.cell(row=row_idx, column=1, value=secteur).font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=BLUE)
    for col, val in enumerate([row["nb_clients"], row["el_normal"], row["el_modere"], row["el_severe"]], 2):
        c = ws.cell(row=row_idx, column=col, value=round(val, 0))
        style(c, bg=bg)

# Ligne total
total_row = 9
ws.cell(row=total_row, column=1, value="TOTAL")
style(ws.cell(row=total_row, column=1), bg=NAVY, bold=True, color=WHITE)
for col in range(2, 6):
    c = ws.cell(row=total_row, column=col)
    c.value = f"=SUM({ws.cell(row=3, column=col).coordinate}:{ws.cell(row=7, column=col).coordinate})"
    style(c, bg=NAVY, bold=True, color=WHITE)

# Graphique
chart = BarChart()
chart.title = "Pertes par Scénario et Secteur"
chart.y_axis.title = "Perte attendue (€)"
chart.x_axis.title = "Secteur"
chart.style = 10
chart.width = 20
chart.height = 12

data = Reference(ws, min_col=3, max_col=5, min_row=2, max_row=7)
cats = Reference(ws, min_col=1, min_row=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "G2")

wb.save("rapport_stress.xlsx")
print("\n✅ Rapport Excel généré : rapport_stress.xlsx")