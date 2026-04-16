import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime, timedelta

np.random.seed(42)

# ── Connexion à la base SQLite ──────────────────────────────
conn = sqlite3.connect("creditsys.db")
cursor = conn.cursor()

# ── Création des 3 tables ───────────────────────────────────
cursor.executescript("""
    DROP TABLE IF EXISTS resultats;
    DROP TABLE IF EXISTS controles;
    DROP TABLE IF EXISTS entites;

    CREATE TABLE entites (
        id          INTEGER PRIMARY KEY,
        nom         TEXT,
        secteur     TEXT,
        responsable TEXT
    );

    CREATE TABLE controles (
        id              INTEGER PRIMARY KEY,
        type_controle   TEXT,
        periodicite     TEXT,
        seuil_alerte    REAL
    );

    CREATE TABLE resultats (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        entite_id   INTEGER,
        controle_id INTEGER,
        date        TEXT,
        statut      TEXT,
        valeur      REAL,
        commentaire TEXT,
        FOREIGN KEY (entite_id)  REFERENCES entites(id),
        FOREIGN KEY (controle_id) REFERENCES controles(id)
    );
""")

conn.commit()
print("✅ Base de données créée : creditsys.db")
print("   Tables : entites, controles, resultats")

# ── Insertion des entités ───────────────────────────────────
entites = [
    (1, "Agence Paris Centre",   "Immobilier", "Dupont Marie"),
    (2, "Agence Lyon Sud",       "Commerce",   "Martin Paul"),
    (3, "Agence Marseille Nord", "Industrie",  "Bernard Luc"),
    (4, "Agence Bordeaux",       "Services",   "Petit Sophie"),
    (5, "Agence Nice Est",       "Agriculture","Moreau Jean"),
]
cursor.executemany("INSERT INTO entites VALUES (?,?,?,?)", entites)

# ── Insertion des types de contrôles ───────────────────────
controles = [
    (1, "Conformité dossiers crédit",  "Mensuel",     0.05),
    (2, "Taux de défaut portefeuille", "Mensuel",     0.15),
    (3, "Délai de traitement",         "Hebdomadaire",48.0),
    (4, "Taux d'endettement moyen",    "Mensuel",     0.45),
    (5, "Incidents de paiement",       "Mensuel",     0.10),
]
cursor.executemany("INSERT INTO controles VALUES (?,?,?,?)", controles)

# ── Génération des résultats sur 6 mois ────────────────────
statuts = ["OK", "OK", "OK", "OK", "ALERTE", "KO"]
resultats = []
date_debut = datetime(2025, 7, 1)

for mois in range(6):
    date = date_debut + timedelta(days=30*mois)
    for entite_id in range(1, 6):
        for controle_id in range(1, 6):
            statut = np.random.choice(statuts, p=[0.6, 0.15, 0.1, 0.08, 0.04, 0.03])
            valeur = round(np.random.uniform(0.01, 0.50), 3)
            resultats.append((
                entite_id, controle_id,
                date.strftime("%Y-%m-%d"),
                statut, valeur, ""
            ))

cursor.executemany(
    "INSERT INTO resultats (entite_id, controle_id, date, statut, valeur, commentaire) VALUES (?,?,?,?,?,?)",
    resultats
)
conn.commit()
print(f"✅ {len(resultats)} résultats de contrôle insérés")

# ── KPI 1 : Taux de conformité global ──────────────────────
print("\n═══ KPI 1 : Taux de conformité global ═══")
query1 = """
    SELECT
        COUNT(*) AS total_controles,
        SUM(CASE WHEN statut='OK' THEN 1 ELSE 0 END) AS nb_ok,
        SUM(CASE WHEN statut='KO' THEN 1 ELSE 0 END) AS nb_ko,
        SUM(CASE WHEN statut='ALERTE' THEN 1 ELSE 0 END) AS nb_alerte,
        ROUND(100.0 * SUM(CASE WHEN statut='OK' THEN 1 ELSE 0 END) / COUNT(*), 1) AS taux_conformite
    FROM resultats;
"""
df_kpi1 = pd.read_sql_query(query1, conn)
print(df_kpi1.to_string(index=False))

# ── KPI 2 : Contrôles KO par entité ────────────────────────
print("\n═══ KPI 2 : Contrôles KO par entité ═══")
query2 = """
    SELECT e.nom, e.secteur,
        COUNT(*) AS nb_ko
    FROM resultats r
    JOIN entites e ON r.entite_id = e.id
    WHERE r.statut = 'KO'
    GROUP BY e.nom, e.secteur
    ORDER BY nb_ko DESC;
"""
df_kpi2 = pd.read_sql_query(query2, conn)
print(df_kpi2.to_string(index=False))

# ── KPI 3 : Tendance mensuelle ──────────────────────────────
print("\n═══ KPI 3 : Tendance mensuelle ═══")
query3 = """
    SELECT strftime('%Y-%m', date) AS mois,
        COUNT(*) AS total,
        ROUND(100.0 * SUM(CASE WHEN statut='OK' THEN 1 ELSE 0 END) / COUNT(*), 1) AS taux_conformite
    FROM resultats
    GROUP BY mois
    ORDER BY mois;
"""
df_kpi3 = pd.read_sql_query(query3, conn)
print(df_kpi3.to_string(index=False))

# ── Rapport Excel automatique ───────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

NAVY  = "1B3A6B"
GREEN = "27AE60"
RED   = "E74C3C"
ORANGE= "E67E22"
WHITE = "FFFFFF"
GREY  = "F2F3F4"

def style(cell, bg=WHITE, bold=False, color="000000", center=True):
    cell.font      = Font(bold=bold, color=color, name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    s = Side(style="thin", color="CCCCCC")
    cell.border    = Border(left=s, right=s, top=s, bottom=s)

# ── Feuille 1 : KPIs globaux ────────────────────────────────
ws1 = wb.active
ws1.title = "KPIs Globaux"
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:E1")
ws1["A1"] = "🏦 CreditSys — Rapport Contrôle Permanent"
style(ws1["A1"], bg=NAVY, bold=True, color=WHITE)
ws1.row_dimensions[1].height = 30

kpis = [
    ("Total contrôles",   150,    GREY),
    ("Contrôles OK",      139,    "D5F5E3"),
    ("Contrôles KO",      2,      "FADBD8"),
    ("Alertes",           9,      "FDEBD0"),
    ("Taux conformité",  "92.7%", "D6E4F0"),
]
for r, (label, val, bg) in enumerate(kpis, 3):
    ws1.row_dimensions[r].height = 25
    c1 = ws1.cell(row=r, column=1, value=label)
    c2 = ws1.cell(row=r, column=2, value=val)
    style(c1, bg=bg, bold=True, color=NAVY, center=False)
    style(c2, bg=bg, bold=True, color=NAVY)
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 15

# ── Feuille 2 : KO par entité ───────────────────────────────
ws2 = wb.create_sheet("KO par Entité")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:C1")
ws2["A1"] = "Contrôles KO par Entité"
style(ws2["A1"], bg=NAVY, bold=True, color=WHITE)

for col, h in enumerate(["Entité", "Secteur", "Nb KO"], 1):
    c = ws2.cell(row=2, column=col, value=h)
    style(c, bg=NAVY, bold=True, color=WHITE)
    ws2.column_dimensions[ws2.cell(row=2, column=col).column_letter].width = 25

for r, row in df_kpi2.iterrows():
    for col, val in enumerate([row["nom"], row["secteur"], row["nb_ko"]], 1):
        c = ws2.cell(row=r+3, column=col, value=val)
        style(c, bg="FADBD8" if row["nb_ko"] > 0 else WHITE)

# ── Feuille 3 : Tendance mensuelle ──────────────────────────
ws3 = wb.create_sheet("Tendance Mensuelle")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:C1")
ws3["A1"] = "Taux de Conformité par Mois"
style(ws3["A1"], bg=NAVY, bold=True, color=WHITE)

for col, h in enumerate(["Mois", "Total contrôles", "Taux conformité (%)"], 1):
    c = ws3.cell(row=2, column=col, value=h)
    style(c, bg=NAVY, bold=True, color=WHITE)
    ws3.column_dimensions[ws3.cell(row=2, column=col).column_letter].width = 22

for r, row in df_kpi3.iterrows():
    taux = row["taux_conformite"]
    bg = "D5F5E3" if taux >= 90 else "FDEBD0" if taux >= 80 else "FADBD8"
    for col, val in enumerate([row["mois"], row["total"], taux], 1):
        c = ws3.cell(row=r+3, column=col, value=val)
        style(c, bg=bg)

wb.save("rapport_controle.xlsx")
print("\n✅ Rapport Excel généré : rapport_controle.xlsx")

conn.close()
print("✅ F2 terminé — base creditsys.db fermée")